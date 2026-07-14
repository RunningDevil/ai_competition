from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from comment_parser import parse_comment_text
from office_common import (
    convert_legacy_office,
    filter_comments,
    fixed_relative_path,
    get_extension,
    make_error,
    normalize_path_text,
    replacement_from_comment,
    source_relative_to_docs,
)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_xlsx_text(path: Path, source_rel: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=False, data_only=True)
    texts: List[Dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            values = [_cell_text(cell.value) for cell in row]
            values = [value for value in values if value]
            if values:
                texts.append({
                    "source": source_rel,
                    "file_type": "xlsx",
                    "location": f"sheet:{sheet.title}:row:{row[0].row if row else 0}",
                    "text": "\t".join(values),
                })
    return texts


def _extract_xlsx_comments(path: Path, source_rel: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=False, data_only=True)
    comments: List[Dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text:
                    location = f"sheet:{sheet.title}:cell:{cell.coordinate}"
                    comments.append(parse_comment_text(cell.comment.text, source_rel, "xlsx", location))
    return comments


def _extract_xls_text(path: Path, source_rel: str) -> List[Dict[str, Any]]:
    import xlrd

    workbook = xlrd.open_workbook(str(path))
    texts: List[Dict[str, Any]] = []
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            values = [_cell_text(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            values = [value for value in values if value]
            if values:
                texts.append({
                    "source": source_rel,
                    "file_type": "xls",
                    "location": f"sheet:{sheet.name}:row:{row_index + 1}",
                    "text": "\t".join(values),
                })
    return texts


def _to_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    is_percent = text.endswith("%")
    text = text.rstrip("%").replace(",", "").replace("，", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100.0 if is_percent else number


def _string_value(value: Any) -> str:
    if value is None:
        return "(空)"
    text = str(value).strip()
    return text if text else "(空)"


def _unique_header(value: Any, index: int, seen: Dict[str, int]) -> str:
    header = str(value or "").strip() or f"列{index + 1}"
    count = seen.get(header, 0)
    seen[header] = count + 1
    return header if count == 0 else f"{header}_{count + 1}"


def _load_xlsx_table(path: Path, logs: List[str]) -> Optional[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = -1
        headers: List[str] = []
        for index, row in enumerate(rows[:50]):
            values = list(row or [])
            non_empty = [value for value in values if _cell_text(value)]
            if len(non_empty) < 2:
                continue
            seen: Dict[str, int] = {}
            headers = [_unique_header(value, col_index, seen) for col_index, value in enumerate(values)]
            header_index = index
            break
        if header_index < 0:
            continue

        records: List[Dict[str, Any]] = []
        for row in rows[header_index + 1 :]:
            values = list(row or [])
            if not any(_cell_text(value) for value in values):
                continue
            record = {headers[col_index]: values[col_index] if col_index < len(values) else None for col_index in range(len(headers))}
            records.append(record)
        if records:
            logs.append(f"Loaded Excel table from sheet {sheet.title}: {len(records)} data rows")
            return {
                "sheet": sheet.title,
                "headers": headers,
                "records": records,
                "header_row": header_index + 1,
            }
    logs.append("No tabular sheet with at least two columns was found")
    return None


def _header_in_title(headers: Iterable[str], title: str, allowed: Iterable[str] | None = None) -> List[str]:
    allowed_set = set(allowed or headers)
    result = []
    for header in headers:
        if header not in allowed_set:
            continue
        if header and header in title:
            result.append(header)
    return result


def _column_profiles(headers: List[str], records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    profiles: Dict[str, Dict[str, Any]] = {}
    for header in headers:
        values = [record.get(header) for record in records if record.get(header) not in (None, "")]
        numeric_values = [_to_number(value) for value in values]
        numeric_count = sum(1 for value in numeric_values if value is not None)
        profiles[header] = {
            "non_empty": len(values),
            "numeric_count": numeric_count,
            "is_numeric": bool(values) and numeric_count >= max(1, int(len(values) * 0.6)),
            "unique_count": len({_string_value(value) for value in values}),
        }
    return profiles


def _infer_aggregate(title: str, value_field: str | None) -> str:
    lower = title.lower()
    if any(word in title for word in ("平均", "均值")) or "average" in lower or "avg" in lower:
        return "avg"
    if any(word in title for word in ("最大", "最高")) or "max" in lower:
        return "max"
    if any(word in title for word in ("最小", "最低")) or "min" in lower:
        return "min"
    if any(word in title for word in ("计数", "条数", "记录数", "个数")) or "count" in lower:
        return "count"
    return "sum" if value_field else "count"


def _infer_chart_type(title: str) -> str:
    lower = title.lower()
    if "pie" in lower or "饼" in title:
        return "pie"
    if "line" in lower or any(word in title for word in ("折线", "趋势")):
        return "line"
    return "bar"


def _infer_pivot_spec(title: str, headers: List[str], records: List[Dict[str, Any]], filters: Dict[str, Any] | None) -> Dict[str, Any]:
    filters = filters or {}
    profiles = _column_profiles(headers, records)
    numeric_headers = [header for header in headers if profiles[header]["is_numeric"]]
    categorical_headers = [header for header in headers if header not in numeric_headers]

    value_field = filters.get("value_field") or filters.get("value") or filters.get("metric")
    row_field = filters.get("row_field") or filters.get("dimension") or filters.get("group_by")
    column_field = filters.get("column_field") or filters.get("column")

    title_numeric = _header_in_title(headers, title, numeric_headers)
    title_categorical = _header_in_title(headers, title, categorical_headers)
    if not value_field and title_numeric:
        value_field = title_numeric[-1]
    if not row_field:
        by_match = re.search(r"按(.+?)(?:统计|汇总|分组|生成|绘制|画|制作|的|$)", title)
        if by_match:
            segment = by_match.group(1)
            for header in categorical_headers:
                if header in segment:
                    row_field = header
                    break
        if not row_field and title_categorical:
            row_field = title_categorical[0]
    if not column_field and len(title_categorical) >= 2:
        for header in title_categorical:
            if header != row_field:
                column_field = header
                break

    if row_field not in headers:
        row_field = categorical_headers[0] if categorical_headers else headers[0]
    if column_field not in headers or column_field == row_field:
        column_field = None
    if value_field not in headers:
        value_field = numeric_headers[0] if numeric_headers else None

    aggregate = str(filters.get("aggregate") or filters.get("agg") or _infer_aggregate(title, value_field)).lower()
    if aggregate not in {"sum", "avg", "count", "max", "min"}:
        aggregate = "sum" if value_field else "count"

    return {
        "row_field": row_field,
        "column_field": column_field,
        "value_field": value_field,
        "aggregate": aggregate,
        "chart_type": str(filters.get("chart_type") or _infer_chart_type(title)).lower(),
    }


def _aggregate(values: List[float], aggregate: str) -> float:
    if aggregate == "count":
        return float(len(values))
    if not values:
        return 0.0
    if aggregate == "avg":
        return sum(values) / len(values)
    if aggregate == "max":
        return max(values)
    if aggregate == "min":
        return min(values)
    return sum(values)


def _clean_number(value: float) -> int | float:
    if abs(value - round(value)) < 0.0000001:
        return int(round(value))
    return round(value, 4)


def _build_pivot(table: Dict[str, Any], spec: Dict[str, Any]) -> Dict[str, Any]:
    row_field = spec["row_field"]
    column_field = spec.get("column_field")
    value_field = spec.get("value_field")
    aggregate = spec["aggregate"]
    buckets: Dict[str, Dict[str, List[float]]] = defaultdict(lambda: defaultdict(list))
    column_values: List[str] = []

    for record in table["records"]:
        row_key = _string_value(record.get(row_field))
        column_key = _string_value(record.get(column_field)) if column_field else "value"
        if column_key not in column_values:
            column_values.append(column_key)
        if aggregate == "count" or not value_field:
            buckets[row_key][column_key].append(1.0)
            continue
        number = _to_number(record.get(value_field))
        if number is not None:
            buckets[row_key][column_key].append(number)

    rows = []
    for row_key in sorted(buckets.keys()):
        values = {column_key: _clean_number(_aggregate(buckets[row_key].get(column_key, []), aggregate)) for column_key in column_values}
        rows.append({"row": row_key, "values": values, "total": _clean_number(sum(float(value) for value in values.values()))})

    return {
        "sheet": table["sheet"],
        "row_field": row_field,
        "column_field": column_field,
        "value_field": value_field,
        "aggregate": aggregate,
        "chart_type": spec["chart_type"],
        "columns": column_values,
        "rows": rows,
    }


def _pivot_relative_path(source_rel: str) -> str:
    normalized = normalize_path_text(source_rel)
    body = normalized[len("docs/") :] if normalized.startswith("docs/") else normalized
    parts = body.split("/")
    file_name = parts[-1]
    stem = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
    parts[-1] = f"{stem}_pivot.xlsx"
    return "output/fixed/" + "/".join(parts)


def _write_pivot_workbook(pivot: Dict[str, Any], source_rel: str, target_abs: Path, logs: List[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "透视汇总"
    sheet["A1"] = "source"
    sheet["B1"] = source_rel
    sheet["A2"] = "sheet"
    sheet["B2"] = pivot["sheet"]
    sheet["A3"] = "row_field"
    sheet["B3"] = pivot["row_field"]
    sheet["C3"] = "value_field"
    sheet["D3"] = pivot.get("value_field") or "count"
    sheet["E3"] = "aggregate"
    sheet["F3"] = pivot["aggregate"]

    header_row = 5
    sheet.cell(header_row, 1, pivot["row_field"])
    columns = pivot["columns"] or ["value"]
    if pivot.get("column_field"):
        for offset, column_name in enumerate(columns, start=2):
            sheet.cell(header_row, offset, column_name)
    else:
        metric_name = pivot.get("value_field") or "count"
        sheet.cell(header_row, 2, f"{pivot['aggregate']}({metric_name})")

    for row_index, row in enumerate(pivot["rows"], start=header_row + 1):
        sheet.cell(row_index, 1, row["row"])
        if pivot.get("column_field"):
            for offset, column_name in enumerate(columns, start=2):
                sheet.cell(row_index, offset, row["values"].get(column_name, 0))
        else:
            sheet.cell(row_index, 2, row["total"])

    max_row = header_row + len(pivot["rows"])
    max_col = 1 + (len(columns) if pivot.get("column_field") else 1)
    if pivot["rows"]:
        chart_type = pivot.get("chart_type") or "bar"
        if chart_type == "pie" and not pivot.get("column_field"):
            chart = PieChart()
        elif chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()
            chart.type = "col"
        chart.title = "透视图"
        if hasattr(chart, "y_axis"):
            chart.y_axis.title = pivot.get("value_field") or "count"
        if hasattr(chart, "x_axis"):
            chart.x_axis.title = pivot["row_field"]
        data = Reference(sheet, min_col=2, max_col=max_col, min_row=header_row, max_row=max_row)
        categories = Reference(sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        sheet.add_chart(chart, "H2")

    target_abs.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target_abs))
    logs.append(f"Excel pivot workbook written: {target_abs}")


def _wants_pivot_output(title: str, filters: Dict[str, Any] | None) -> bool:
    filters = filters or {}
    if filters.get("make_pivot") or filters.get("chart_type") or filters.get("row_field") or filters.get("value_field"):
        return True
    lower = title.lower()
    return any(word in title for word in ("透视", "透视图", "数据透视", "图表", "柱状图", "折线图", "饼图")) or "pivot" in lower


def _build_pivot_output(path: Path, source_rel: str, wiki_root: str | Path, payload: Dict[str, Any], logs: List[str]) -> Optional[Dict[str, Any]]:
    title = str(payload.get("question_title") or "")
    filters = payload.get("filters") or {}
    table_path = path
    if path.suffix.lower() == ".xls":
        converted = convert_legacy_office(path, logs)
        if not converted or converted.suffix.lower() != ".xlsx":
            logs.append("Legacy .xls could not be converted for pivot output")
            return None
        table_path = converted

    table = _load_xlsx_table(table_path, logs)
    if not table:
        return None
    spec = _infer_pivot_spec(title, table["headers"], table["records"], filters)
    pivot = _build_pivot(table, spec)
    if not pivot["rows"]:
        logs.append("Pivot aggregation produced no rows")
        return None

    target_rel = _pivot_relative_path(source_rel)
    target_abs = Path(wiki_root) / target_rel
    _write_pivot_workbook(pivot, source_rel, target_abs, logs)
    return {
        "answer": {"source": source_rel, "target": target_rel},
        "fixed_files": [target_rel],
        "pivot": pivot,
        "tables": [
            {
                "source": source_rel,
                "sheet": pivot["sheet"],
                "kind": "pivot",
                "summary": (
                    f"透视表: 按 {pivot['row_field']} "
                    f"{'和 ' + pivot['column_field'] + ' ' if pivot.get('column_field') else ''}"
                    f"统计 {pivot['aggregate']}({pivot.get('value_field') or 'count'}), "
                    f"生成 {target_rel}"
                ),
                "row_field": pivot["row_field"],
                "column_field": pivot.get("column_field"),
                "value_field": pivot.get("value_field"),
                "aggregate": pivot["aggregate"],
                "target": target_rel,
            }
        ],
    }


def extract_text(path: Path, wiki_root: str | Path, logs: List[str]) -> List[Dict[str, Any]]:
    ext = get_extension(path)
    source_rel = source_relative_to_docs(path, wiki_root)
    if ext == ".xlsx":
        return _extract_xlsx_text(path, source_rel)
    if ext == ".xls":
        converted = convert_legacy_office(path, logs)
        if converted and converted.suffix.lower() == ".xlsx":
            return _extract_xlsx_text(converted, source_rel)
        try:
            return _extract_xls_text(path, source_rel)
        except Exception as exc:
            logs.append(f"xlrd fallback failed for .xls: {exc}")
    return []


def extract_comments(path: Path, wiki_root: str | Path, logs: List[str]) -> List[Dict[str, Any]]:
    ext = get_extension(path)
    source_rel = source_relative_to_docs(path, wiki_root)
    if ext == ".xlsx":
        return _extract_xlsx_comments(path, source_rel)
    if ext == ".xls":
        converted = convert_legacy_office(path, logs)
        if converted and converted.suffix.lower() == ".xlsx":
            return _extract_xlsx_comments(converted, source_rel)
        logs.append("Legacy .xls comments could not be reliably extracted without conversion")
    return []


def analyze(path: Path, wiki_root: str | Path, logs: List[str], payload: Dict[str, Any] | None = None) -> Dict[str, Any]:
    payload = payload or {}
    texts = extract_text(path, wiki_root, logs)
    source_rel = source_relative_to_docs(path, wiki_root)
    if get_extension(path) in {".xlsx", ".xls"} and _wants_pivot_output(str(payload.get("question_title") or ""), payload.get("filters") or {}):
        pivot_result = _build_pivot_output(path, source_rel, wiki_root, payload, logs)
        if pivot_result:
            return {
                "status": "ok",
                "task_type": "excel_analyze",
                "texts": texts,
                "comments": [],
                "answer": pivot_result["answer"],
                "fixed_files": pivot_result["fixed_files"],
                "tables": pivot_result["tables"],
                "pivot": pivot_result["pivot"],
                "logs": logs,
            }

    sheet_rows: Dict[str, int] = {}
    for item in texts:
        location = item.get("location", "")
        if location.startswith("sheet:"):
            parts = location.split(":")
            if len(parts) >= 2:
                sheet_rows[parts[1]] = sheet_rows.get(parts[1], 0) + 1
    summary = [f"{sheet}: {rows} non-empty rows" for sheet, rows in sheet_rows.items()]
    return {
        "status": "ok",
        "task_type": "excel_analyze",
        "texts": texts,
        "comments": [],
        "answer": {"datas": summary or [f"{source_rel}: no non-empty rows found"]},
        "fixed_files": [],
        "tables": [{"source": source_rel, "sheet": sheet, "non_empty_rows": rows} for sheet, rows in sheet_rows.items()],
        "logs": logs,
    }


def _write_replaced_xlsx(source: Path, target: Path, old: str, new: str, logs: List[str]) -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(str(source), read_only=False, data_only=False)
    replaced_count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if value.startswith("="):
                    continue
                if old not in value:
                    continue
                cell.value = value.replace(old, new, 1)
                replaced_count += 1
                break
            if replaced_count:
                break
        if replaced_count:
            break
    if replaced_count:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(target))
        logs.append(f"Replaced Excel text: {old} -> {new}")
    else:
        logs.append(f"Replacement text not found in Excel cells: {old}")
    return replaced_count


def conservative_fix(path: Path, wiki_root: str | Path, logs: List[str], filters: Dict[str, Any] | None = None) -> Dict[str, Any]:
    if get_extension(path) != ".xlsx":
        return make_error("fix_comments", f"Excel repair only supports .xlsx for now: {normalize_path_text(str(path))}", logs)
    comments = extract_comments(path, wiki_root, logs)
    if filters:
        comments = filter_comments(comments, filters)
    if not comments:
        return make_error("fix_comments", f"No reliable Excel comments found for {normalize_path_text(str(path))}", logs)
    source_rel = source_relative_to_docs(path, wiki_root)
    target_rel = fixed_relative_path(source_rel)
    target_abs = Path(wiki_root) / target_rel
    for comment in comments:
        replacement = replacement_from_comment(comment)
        if not replacement:
            continue
        old, new = replacement
        replaced_count = _write_replaced_xlsx(path, target_abs, old, new, logs)
        if replaced_count:
            return {
                "status": "ok",
                "task_type": "fix_comments",
                "texts": [],
                "comments": comments,
                "answer": {"source": source_rel, "target": target_rel},
                "fixed_files": [target_rel],
                "logs": logs,
            }
    logs.append("Excel comments found, but no deterministic replacement instruction could be applied")
    return make_error("fix_comments", f"No reliable Excel repair rule implemented for {normalize_path_text(str(path))}", logs)
